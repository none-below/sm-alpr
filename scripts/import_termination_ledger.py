#!/usr/bin/env python3
"""Import an external "Flock terminated/paused contracts" spreadsheet.

The activist tracker we're consuming is shared as an .xlsx workbook with
two sheets (Flock_ALPR_Agencies, Timeline_View). It tracks U.S. agencies
that have terminated or paused Flock ALPR contracts, plus citations.

What we want from it:
  1. The citation URLs, fed into the article-fetch queue so the article
     registry grows. The curator picks tags (outcome:terminated etc.)
     from article content using the existing controlled vocab — no
     bespoke tagging path for this import.
  2. A small audit trail of what we imported and when
     (assets/termination_ledger.json), so re-imports can detect upstream
     changes via file hash. This is a bootstrap, not a parallel data
     store — long-term, the canonical truth is article_registry +
     agency_registry annotations, not this file. Downstream consumers
     (report, sharing map) should NOT query this file; they should
     query the registries with generic tag filters that work for any
     article, regardless of how it was discovered.

The xlsx itself is not committed — it's an external activist's doc; we
only commit our derivation. Re-import is idempotent: same file produces
the same ledger; queue-add dedupes new URLs against the existing queue
and registry.

Usage:
  uv run python scripts/import_termination_ledger.py PATH/TO/FILE.xlsx
  uv run python scripts/import_termination_ledger.py FILE.xlsx --add-domains
  uv run python scripts/import_termination_ledger.py FILE.xlsx --add-domains --queue
  uv run python scripts/import_termination_ledger.py FILE.xlsx --dry-run

Flags:
  --add-domains  Append unknown domains to assets/sources.json at tier 2
                 (or tier 1 for an inline allowlist of papers/NPR member
                 stations) with stance=unknown, last_reviewed=today.
  --queue        Forward queueable URLs to scripts/article_queue_add.py.
  --dry-run      Compute everything; write nothing.
"""

import argparse
import functools
import hashlib
import json
import re
import subprocess
import sys
from datetime import date, datetime, timezone
from pathlib import Path
from urllib.parse import urlparse

import openpyxl

print = functools.partial(print, flush=True)

ROOT = Path(__file__).resolve().parent.parent
SOURCES_PATH = ROOT / "assets" / "sources.json"
LEDGER_PATH = ROOT / "assets" / "termination_ledger.json"
QUEUE_ADD = ROOT / "scripts" / "article_queue_add.py"

# Domains we promote to tier 1 on auto-add: established regional papers,
# NPR member stations, and major-market network O&Os. Local TV affiliates,
# weeklies, government press releases, and uncategorized sites stay at
# tier 2 (subagent-jail per assets/sources.json _handler_policy). The
# distinction matters because tier 1 clean reads happen in-context.
TIER_1_AUTO_ADD = {
    "cbsnews.com",
    "denvergazette.com",
    "denverpost.com",
    "freep.com",        # Detroit Free Press
    "heraldnet.com",    # Everett Herald
    "klcc.org",         # NPR Eugene
    "kut.org",          # NPR Austin
    "madison.com",      # Wisconsin State Journal
    "mlive.com",        # MLive Michigan
    "newsobserver.com", # Raleigh News & Observer
    "northcountrypublicradio.org",
    "opb.org",          # Oregon Public Broadcasting
    "seattletimes.com",
    "thenewstribune.com",  # Tacoma News Tribune
    "theolympian.com",
    "timesunion.com",   # Albany Times Union
    "toledoblade.com",
    "wskg.org",         # NPR Binghamton
}


def normalize_domain(host: str) -> str:
    h = (host or "").lower()
    if h.startswith("www."):
        h = h[4:]
    return h


def domain_in_allowlist(host: str, allowed: set[str]) -> str | None:
    h = normalize_domain(host)
    if h in allowed:
        return h
    parts = h.split(".")
    for i in range(1, len(parts) - 1):
        c = ".".join(parts[i:])
        if c in allowed:
            return c
    return None


def parse_date(value) -> str | None:
    """Sheet 1 has datetimes; Sheet 2 has free-text ('Sept 2025', '2023',
    'Oct–Dec 2025'). Coerce to ISO-ish string; pass through anything we
    can't parse so the human-readable form survives."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def split_urls(cell) -> list[str]:
    """The Source URLs cell mixes whitespace, semicolons, and commas as
    separators. Strip trailing punctuation (periods on sentence-end URLs)."""
    if not cell:
        return []
    out = []
    for tok in re.split(r"[\s;,]+", str(cell)):
        tok = tok.strip().rstrip(".")
        if tok.startswith("http"):
            # Drop fragment; keep query (article URLs sometimes need ?id=...)
            p = urlparse(tok)
            if not p.netloc:
                continue
            out.append(p._replace(fragment="").geturl())
    return out


def read_workbook(path: Path) -> tuple[list[dict], str]:
    """Return (action records, sha256 of file).

    Each sheet contributes its own records; we don't dedupe across sheets
    because the agency_name encoding differs (Sheet 1: 'Sedona', Sheet 2:
    'City of Sedona / Sedona PD'). URL-level dedup happens at queue time.
    """
    sha = hashlib.sha256(path.read_bytes()).hexdigest()
    wb = openpyxl.load_workbook(path, data_only=True)
    actions = []
    expected = ["Agency Name", "Agency Type", "State", "Action Type",
                "Approx. Date", "Notes", "Source URLs"]
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [h for h in rows[0] if h is not None]
        # Tolerate trailing None columns; fail loudly on schema drift.
        for col in expected:
            if col not in headers:
                print(f"WARN: sheet {sn!r} missing column {col!r}; skipping",
                      file=sys.stderr)
                break
        else:
            idx = {c: headers.index(c) for c in expected}
            for row in rows[1:]:
                agency = row[idx["Agency Name"]]
                if not agency:
                    continue
                actions.append({
                    "agency_name": str(agency).strip(),
                    "agency_type": (row[idx["Agency Type"]] or "").strip() or None,
                    "state": (row[idx["State"]] or "").strip() or None,
                    "action_type": (row[idx["Action Type"]] or "").strip() or None,
                    "approx_date": parse_date(row[idx["Approx. Date"]]),
                    "notes": (str(row[idx["Notes"]]).strip()
                              if row[idx["Notes"]] is not None else None),
                    "source_urls": split_urls(row[idx["Source URLs"]]),
                    "sheet": sn,
                })
    return actions, sha


def collect_unique_urls(actions: list[dict]) -> list[tuple[str, str]]:
    """Return list of (url, normalized_apex_domain), order preserved,
    duplicates collapsed."""
    seen = set()
    out = []
    for a in actions:
        for u in a["source_urls"]:
            if u in seen:
                continue
            seen.add(u)
            host = urlparse(u).hostname or ""
            out.append((u, normalize_domain(host)))
    return out


def load_sources() -> dict:
    return json.loads(SOURCES_PATH.read_text())


def add_domains_to_sources(sources: dict, new_domains: list[str],
                           today: str) -> int:
    """Append entries for unknown domains. Returns count added."""
    existing = {s["domain"].lower() for s in sources["sources"]}
    added = 0
    for d in new_domains:
        if d in existing:
            continue
        tier = 1 if d in TIER_1_AUTO_ADD else 2
        sources["sources"].append({
            "domain": d,
            "tier": tier,
            "name": d,  # placeholder; human can rename in review
            "stance": "unknown",
            "last_reviewed": today,
            "notes": "auto-added from termination_ledger import; review tier/stance",
        })
        existing.add(d)
        added += 1
    # Keep the file deterministic: sort by domain after our schema/header keys.
    sources["sources"].sort(key=lambda s: s["domain"])
    return added


def write_ledger(actions: list[dict], src_sha: str, src_filename: str,
                 dry_run: bool) -> None:
    # Preserve imported_at when the source hash and action list are unchanged
    # — otherwise re-running on the same xlsx churns the file on disk and
    # produces a noisy git diff. Only the timestamp updates when content
    # actually changes.
    prior_imported_at = None
    if LEDGER_PATH.exists():
        try:
            prior = json.loads(LEDGER_PATH.read_text())
            if (prior.get("source_file_sha256") == src_sha
                    and prior.get("actions") == actions):
                prior_imported_at = prior.get("imported_at")
        except (json.JSONDecodeError, OSError):
            pass

    payload = {
        "_schema": {
            "imported_at": "ISO-8601 UTC of last import that changed content",
            "source_file": "filename of upstream xlsx (not committed)",
            "source_file_sha256": "sha256 of upstream xlsx at import time",
            "actions": "list of agency-action records harvested from sheets",
        },
        "_note": (
            "Bootstrap import audit trail, not a canonical data store. "
            "Derived from an external activist tracker of Flock terminated/"
            "paused contracts; the source xlsx is not committed. The "
            "canonical record of contract-termination evidence is "
            "assets/article_registry.json (with generic outcome:* tags "
            "applied by the curator). Re-import via "
            "scripts/import_termination_ledger.py."
        ),
        "imported_at": prior_imported_at or datetime.now(timezone.utc).strftime(
            "%Y-%m-%dT%H:%M:%SZ"),
        "source_file": src_filename,
        "source_file_sha256": src_sha,
        "actions": actions,
    }
    if dry_run:
        print(f"[dry-run] would write {LEDGER_PATH} ({len(actions)} actions)")
        return
    LEDGER_PATH.write_text(json.dumps(payload, indent=2) + "\n")
    if prior_imported_at:
        print(f"ledger unchanged ({len(actions)} actions); kept imported_at={prior_imported_at}")
    else:
        print(f"wrote {LEDGER_PATH} ({len(actions)} actions)")


def queue_urls(urls: list[str], dry_run: bool) -> None:
    if not urls:
        return
    if dry_run:
        print(f"[dry-run] would enqueue {len(urls)} URL(s) via {QUEUE_ADD.name}")
        return
    # No --tags-hint — the curator derives tags from article content via
    # the existing controlled vocab. Hinting here would create a tagging
    # path specific to this import, which is the opposite of what we want.
    cmd = [sys.executable, str(QUEUE_ADD),
           "--discovered-by", "termination-ledger-import",
           *urls]
    rc = subprocess.run(cmd).returncode
    if rc not in (0, 1):
        # 1 = some URLs rejected (we should have filtered, but be defensive)
        print(f"WARN: queue script exited {rc}", file=sys.stderr)


def main() -> int:
    p = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("xlsx", help="path to the upstream .xlsx workbook")
    p.add_argument("--add-domains", action="store_true",
                   help="append unknown domains to assets/sources.json")
    p.add_argument("--queue", action="store_true",
                   help="forward queueable URLs to article_queue_add.py")
    p.add_argument("--dry-run", action="store_true",
                   help="compute everything; write nothing")
    args = p.parse_args()

    src = Path(args.xlsx).expanduser()
    if not src.is_file():
        print(f"ERROR: {src} not found", file=sys.stderr)
        return 3

    actions, sha = read_workbook(src)
    print(f"read {len(actions)} action record(s) from {src.name} (sha256={sha[:12]}…)")

    urls = collect_unique_urls(actions)
    print(f"unique URLs: {len(urls)}")

    sources = load_sources()
    allowed = {s["domain"].lower() for s in sources["sources"]}

    queueable: list[str] = []
    blocked_by_domain: dict[str, list[str]] = {}
    for u, dom in urls:
        if domain_in_allowlist(dom, allowed):
            queueable.append(u)
        else:
            blocked_by_domain.setdefault(dom, []).append(u)

    print(f"  queueable (domain already in sources.json): {len(queueable)}")
    print(f"  blocked (domain not in sources.json): "
          f"{sum(len(v) for v in blocked_by_domain.values())} URL(s) "
          f"across {len(blocked_by_domain)} domain(s)")

    if args.add_domains and blocked_by_domain:
        today = date.today().isoformat()
        added = add_domains_to_sources(
            sources, sorted(blocked_by_domain.keys()), today)
        if args.dry_run:
            print(f"[dry-run] would add {added} domain(s) to sources.json")
        else:
            SOURCES_PATH.write_text(json.dumps(sources, indent=2) + "\n")
            print(f"added {added} domain(s) to {SOURCES_PATH.name}")
        # Recompute queueable after expansion.
        allowed = {s["domain"].lower() for s in sources["sources"]}
        queueable = [u for u, dom in urls
                     if domain_in_allowlist(dom, allowed)]
        print(f"  queueable after --add-domains: {len(queueable)}")
    elif blocked_by_domain:
        # Help the human scan: domains with their article counts.
        print("\nblocked domains (run with --add-domains to bulk-add):")
        for d in sorted(blocked_by_domain.keys()):
            n = len(blocked_by_domain[d])
            tier = 1 if d in TIER_1_AUTO_ADD else 2
            print(f"  {d:40s}  {n} article(s)  → would auto-add tier {tier}")

    write_ledger(actions, sha, src.name, args.dry_run)

    if args.queue:
        queue_urls(queueable, args.dry_run)
    elif queueable:
        print(f"\n{len(queueable)} URL(s) ready to queue. Re-run with --queue.")

    return 0


if __name__ == "__main__":
    sys.exit(main())
