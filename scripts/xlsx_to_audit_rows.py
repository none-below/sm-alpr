#!/usr/bin/env python3
# SPDX-License-Identifier: AGPL-3.0-or-later
# SPDX-FileCopyrightText: 2026 zero-below
"""
Convert Flock "Network Audit" .xlsx exports into a PRA-import audit source
(`pra-<id>.json`) under assets/transparency.flocksafety.com/<slug>/, matching
the shape scripts/import_pra_audit.py writes. build_audit_log.py then unions it
into docs/data/audit/<slug>.json, which feeds the justifications page — so an
agency whose audit arrived as spreadsheets (not the usual scraped-portal CSV or
PRA PDF) flows through the same pipeline and deploys with the site.

The committed source carries only what the audit pipeline needs — id, searchDate,
networkCount, reason. The searcher is redacted to "***" by default (matching the
SMPD import; no downstream consumer uses it, and it avoids publishing individual
officers' names in a derived build artifact); pass --keep-user to retain it.

The justifications page is per-agency: it shows an agency's OWN searches. Flock's
**Organization Audit** is exactly that, so prefer it as the source. A **Network
Audit** instead lists every search that touched the target agency's network —
run by the agency AND its sharing partners — so it needs `--org` to filter down
to the one searching agency before it means the same thing. `id` is synthesized
(sha1 of the stable row content) since the reason-bearing releases carry no
export UUID; rows dedupe by that id and sort by (searchDate, id).

Input is either the raw workbooks or the committed `*.ndjson.gz` dumps written
by scripts/xlsx_to_audit_ndjson.py. Prefer the NDJSON: it is in the repo, so the
justifications rebuild without the multi-hundred-MB spreadsheets on disk, and it
already carries that script's header repairs.

The raw .xlsx are NOT needed at build time — only this committed json is. Re-run
this when a new PRA release arrives; CI reads the committed json.

Usage:
    uv run python scripts/xlsx_to_audit_rows.py \
        --slug redwood-city-ca-pd --pra-id 26-217 --org "Redwood City CA PD" \
        assets/redwood-city-pras/"PRA 26-217 1st Release.xlsx" ...

    uv run python scripts/xlsx_to_audit_rows.py \
        --slug los-altos-ca-pd --pra-id 26-366 --org "Los Altos CA PD" \
        --reason-required \
        assets/los-altos-pras/json/Los_Altos_PD_Organizational_Audit_*.ndjson.gz
"""
import argparse
import gzip
import hashlib
import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

PORTAL_ROOT = Path("assets/transparency.flocksafety.com")
# Agencies mark withheld cells either with Flock's native `***` or by typing over
# the column in Excel (Los Altos' 2025 workbooks); neither is a real value.
BLANK = {"", "***", "REDACTED"}
MAIN_HINTS = {"Search Time", "Org Name", "Search Type"}


def parse_iso(v):
    if not isinstance(v, str):
        return None
    s = v.strip()
    for fmt in ("%m/%d/%Y, %I:%M:%S %p UTC", "%m/%d/%Y, %I:%M:%S %p", "%m/%d/%Y %I:%M:%S %p UTC"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%dT%H:%M:%SZ")
        except ValueError:
            continue
    return None


def clean(v):
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s in BLANK else s


def find_header(ws):
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        hdr = [str(c).strip() if c is not None else "" for c in row]
        if MAIN_HINTS & set(hdr):
            return {h: i for i, h in enumerate(hdr)}
    return None


def iter_xlsx(fp):
    """Yield one dict per data row, for every sheet that looks like an audit.

    Agencies ship analysis tabs beside the export (Redwood City's releases carry
    "SearchReason"/"Searcher" pivots) and month-per-sheet workbooks (Los Altos),
    so sheets are selected by header rather than by position.
    """
    wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    seen_header = False
    for ws in wb.worksheets:
        idx = find_header(ws)
        if not idx or "Search Time" not in idx:
            continue
        seen_header = True
        by_pos = {i: h for h, i in idx.items()}
        rows = ws.iter_rows(values_only=True)
        next(rows, None)
        for r in rows:
            yield {by_pos[i]: v for i, v in enumerate(r) if i in by_pos}
    wb.close()
    if not seen_header:
        print(f"  skip {fp.name}: no recognizable header", file=sys.stderr)


def iter_ndjson(fp):
    """Yield one dict per line of a (gzipped) NDJSON audit dump."""
    opener = gzip.open if fp.suffix == ".gz" else open
    with opener(fp, "rt", encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if line:
                yield json.loads(line)


def iter_records(fp):
    """Column-name-keyed rows from either a workbook or a committed NDJSON dump.

    The NDJSON path lets an agency's justifications rebuild from data committed
    to the repo (assets/<agency>-pras/json/), with the header repairs from
    scripts/xlsx_to_audit_ndjson.py already applied, instead of needing the raw
    multi-hundred-MB spreadsheets on disk.
    """
    if fp.name.endswith(".ndjson.gz") or fp.suffix == ".ndjson":
        return iter_ndjson(fp)
    return iter_xlsx(fp)


def convert(files, org_filter, reason_required, keep_user):
    by_id = {}
    contributing = 0
    org_lower = org_filter.lower() if org_filter else None
    for fp in files:
        added = 0
        for rec in iter_records(fp):
            if org_lower is not None:
                if org_lower not in clean(rec.get("Org Name")).lower():
                    continue
            reason = clean(rec.get("Reason"))
            if reason_required and not reason:
                continue
            search_date = parse_iso(clean(rec.get("Search Time")))
            if not search_date:
                continue
            user = clean(rec.get("Name"))
            nc = clean(rec.get("Total Networks Searched"))
            case = clean(rec.get("Case #"))
            rid = hashlib.sha1("|".join([user, search_date, nc, reason, case]).encode("utf-8")).hexdigest()[:16]
            if rid in by_id:
                continue
            row = {
                "id": rid,
                "userId": user if (keep_user and user) else "***",
                "searchDate": search_date,
                "networkCount": nc or "0",
            }
            if reason:
                row["reason"] = reason
            if case:
                row["caseNumber"] = case
            by_id[rid] = row
            added += 1
        if added:
            contributing += 1
        print(f"  {fp.name}: +{added} rows", file=sys.stderr)
    return by_id, contributing


def main():
    ap = argparse.ArgumentParser(description="Convert Flock network-audit xlsx to a PRA-import audit source")
    ap.add_argument("files", nargs="+", help="xlsx files to read")
    ap.add_argument("--slug", required=True, help="agency slug under assets/transparency.flocksafety.com/")
    ap.add_argument("--pra-id", required=True, help="PRA request id (output file is pra-<id>.json)")
    ap.add_argument("--org", default=None, help="filter to rows whose Org Name contains this (the searching agency)")
    ap.add_argument("--reason-required", action="store_true",
                    help="keep only rows with a non-blank reason (recommended for justifications)")
    ap.add_argument("--keep-user", action="store_true",
                    help="keep the searcher name instead of redacting it to '***'")
    ap.add_argument("--out", default=None, help="output path (default the portal folder's pra-<id>.json)")
    args = ap.parse_args()

    files = [Path(f) for f in args.files]
    missing = [f for f in files if not f.exists()]
    if missing:
        print(f"missing files: {missing}", file=sys.stderr)
        return 1

    portal_dir = PORTAL_ROOT / args.slug
    if args.out is None and not portal_dir.is_dir():
        print(f"portal folder does not exist: {portal_dir}", file=sys.stderr)
        return 1

    by_id, contributing = convert(files, args.org, args.reason_required, args.keep_user)
    if not by_id:
        print("no rows produced", file=sys.stderr)
        return 1

    rows = sorted(by_id.values(), key=lambda r: (r["searchDate"], r["id"]))
    dates = [r["searchDate"][:10] for r in rows]
    payload = {
        "source": f"Flock network-audit xlsx, PRA #{args.pra_id}, converted by scripts/xlsx_to_audit_rows.py",
        "integrity": {
            "org_filter": args.org,
            "source_files": contributing,
            "row_count": len(rows),
            "search_date_min": min(dates),
            "search_date_max": max(dates),
        },
        "search_audit_csv": rows,
    }
    out = Path(args.out) if args.out else portal_dir / f"pra-{args.pra_id}.json"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(payload, indent=2) + "\n")
    print(f"wrote {out} ({len(rows)} rows, {min(dates)}..{max(dates)})", file=sys.stderr)
    print("rebuild:  uv run python scripts/build_audit_log.py --portal " + args.slug, file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
